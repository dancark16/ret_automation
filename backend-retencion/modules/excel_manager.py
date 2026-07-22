from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


HEADERS = ["FECHA", "CLIENTE", "FACTURA", "No. RETENCIÓN", "RENTA %", "IVA %", "TOTAL RET.", "OBSERVACIÓN"]
YELLOW = PatternFill("solid", fgColor="FFFF00")


def _get_or_create_sheet(path: Path) -> tuple:
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Retenciones"
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 22
    return wb, ws


def register_retention(path: Path, data, invoice_date: str, observation: str = "") -> int:
    wb, ws = _get_or_create_sheet(path)

    # Convertir fecha de factura "17/07/2026" → datetime
    try:
        fecha = datetime.strptime(invoice_date, "%d/%m/%Y")
        fecha_str = fecha.strftime("%-d-%b").upper()  # "17-JUL" — Linux
    except Exception:
        try:
            fecha_str = datetime.strptime(invoice_date, "%d/%m/%Y").strftime("%d-%b").lstrip("0")
        except Exception:
            fecha_str = invoice_date

    row = [
        fecha_str,
        data.client_name,
        data.invoice_sequential,
        data.ret_number,
        data.renta_pct,
        data.iva_pct,
        round(data.renta_value + data.iva_value, 2),
        observation,
    ]

    next_row = ws.max_row + 1
    for col, value in enumerate(row, 1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = Alignment(horizontal="left")
        if observation and col == len(HEADERS):
            cell.fill = YELLOW

    wb.save(path)
    return next_row