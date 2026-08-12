import time
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

YELLOW = PatternFill("solid", fgColor="FFFF00")
BLUE_FILL = PatternFill("solid", fgColor="003DA5")
WHITE_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
NORMAL_FONT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="D0DBF0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = ["FECHA", "CLIENTE", "FACTURA", "No. RETENCIÓN", "RENTA %", "IVA %", "TOTAL RET.", "OBSERVACIÓN"]
COL_WIDTHS = [10, 28, 10, 24, 10, 10, 13, 42]
SHEET = "RETENCIONES"

MONTHS_ES = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
    7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic"
}


def _create_workbook(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET

    # Título
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "RETENCIONES PROCESADAS"
    title.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    title.fill = PatternFill("solid", fgColor="002880")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers fila 2
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = WHITE_FONT
        cell.fill = BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    # Tabla Excel (empieza en fila 2)
    table = Table(displayName="TablaRetenciones", ref=f"A2:H2")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(path)


def _next_empty_row(ws) -> int:
    row = 3
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    return row


def _expand_table(ws, last_row: int):
    """Expande el rango de la tabla para incluir la nueva fila."""
    for tbl in ws.tables.values():
        tbl.ref = f"A2:H{last_row}"


def register_retention(path: Path, job: dict, invoice_date: str, observation: str = ""):
    if not path.exists():
        _create_workbook(path)

    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active

    try:
        fecha = datetime.strptime(invoice_date, "%d/%m/%Y")
        fecha_str = f"{fecha.day}-{MONTHS_ES[fecha.month]}"
    except Exception:
        fecha_str = invoice_date

    from client_aliases import resolve_client
    total_ret = round(job.get("renta_value", 0) + job.get("iva_value", 0), 2)

    row_data = [
        fecha_str,
        resolve_client(job.get("client_name", "")),
        job.get("invoice_sequential", ""),
        job.get("ret_number", ""),
        job.get("renta_pct", 0),
        job.get("iva_pct", 0),
        total_ret,
        observation,
    ]

    next_row = _next_empty_row(ws)
    fill = YELLOW if observation else None

    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
        if fill:
            cell.fill = fill

    ws.row_dimensions[next_row].height = 18
    _expand_table(ws, next_row)

    for attempt in range(5):
        try:
            wb.save(path)
            return
        except PermissionError:
            if attempt < 4:
                print(f"[EXCEL] Archivo bloqueado, ciérralo... reintentando en 5s ({attempt+1}/5)")
                time.sleep(5)
            else:
                raise
